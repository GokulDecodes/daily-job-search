"""
=============================================================
  GOKUL'S DAILY JOB SEARCH — LIVE VERSION (Adzuna API)
  Fetches fresh Angular/TypeScript/React Native jobs daily
  Resolves exact company job page URLs automatically
=============================================================
"""

import os
import csv
import json
import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# ─────────────────────────────────────────────
# SECTION 1: CONFIG — YOUR PROFILE
# ─────────────────────────────────────────────

APP_ID  = os.environ.get("ADZUNA_APP_ID", "")
APP_KEY = os.environ.get("ADZUNA_APP_KEY", "")

PROFILE = {
    "name": "Gokul P",
    "experience_years": 3,
    "primary_skills": ["Angular", "TypeScript", "React Native", "RxJS"],
    "all_skills": [
        "Angular", "TypeScript", "JavaScript", "React", "React Native",
        "RxJS", "HTML5", "CSS3", "SCSS", "Node.js", "REST APIs",
        "TestCafe", "Cypress", "Selenium", "Jasmine", "Jest", "Karma",
        "Jenkins", "GitLab CI/CD", "GitHub Actions", "PostgreSQL",
        "MySQL", "Ionic", "Expo", "Bootstrap", "Webpack", "Figma",
        "JWT", "OAuth2", "SOLID", "Agile", "JIRA"
    ],
}

SEARCH_QUERIES = [
    "Angular TypeScript developer Chennai",        # 1
    "Angular TypeScript developer Bangalore",      # 2
    "React Native developer Chennai",              # 3
    "React Native developer Bangalore",            # 4

    # ── BY JOB TITLE + LEVEL ────────────────────
    "junior Angular developer Chennai",            # 5
    "junior frontend developer Bangalore",         # 6
    "associate software engineer frontend",        # 7
    "fresher Angular TypeScript developer",        # 8

    # ── BY DOMAIN ───────────────────────────────
    "Angular developer product startup India",     # 9
    "frontend developer IT services Chennai",      # 10
    "Angular developer industrial automation",     # 11
    "UI developer TypeScript Chennai",             # 12

    # ── BROAD CATCH-ALL ─────────────────────────
    "frontend engineer 1 year experience India",   # 13
    "Angular RxJS developer India",                # 14
    "React Native Expo developer India",           # 15
    "frontend developer Bangalore 2 years",        # 16
]

HISTORY_FILE = "seen_jobs.json"

# ─────────────────────────────────────────────
# SECTION 2: FETCH JOBS FROM ADZUNA API
# ─────────────────────────────────────────────

def fetch_jobs(query: str, pages: int = 2) -> list:
    all_results = []
    base_url = "https://api.adzuna.com/v1/api/jobs/in/search"

    for page in range(1, pages + 1):
        try:
            params = {
                "app_id": APP_ID,
                "app_key": APP_KEY,
                "results_per_page": 20,
                "what": query,
                "where": "india",
                "content-type": "application/json",
                "page": page,
                "sort_by": "date",
                "max_days_old": 2,
            }
            response = requests.get(base_url, params=params, timeout=15)
            if response.status_code == 200:
                data = response.json()
                results = data.get("results", [])
                all_results.extend(results)
                if len(results) < 20:
                    break
            else:
                print(f"  ⚠️  API error {response.status_code} for: {query}")
                break
        except Exception as e:
            print(f"  ⚠️  Request failed: {e}")
            break

    return all_results

# ─────────────────────────────────────────────
# SECTION 3: RESOLVE EXACT JOB URL
# Follows Adzuna redirect to get real company URL
# ─────────────────────────────────────────────

def resolve_exact_url(redirect_url: str) -> str:
    """
    Adzuna gives a redirect_url like:
      https://adzuna.in/jobs/land/ad/12345?...
    This function follows the redirect to get the
    EXACT company job page URL e.g.:
      https://careers.zoho.com/job/12345
      https://jobs.lever.co/company/abc
      https://naukri.com/job-listings-xyz
    Falls back to Adzuna link if resolve fails.
    """
    if not redirect_url:
        return ""
    try:
        response = requests.get(
            redirect_url,
            allow_redirects=True,
            timeout=10,
            headers={
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
            }
        )
        final_url = response.url
        # If it stayed on Adzuna, return original redirect as fallback
        if "adzuna.com" in final_url or "adzuna.in" in final_url:
            return redirect_url
        return final_url
    except Exception:
        return redirect_url  # fallback

# ─────────────────────────────────────────────
# SECTION 4: FILTER & PARSE
# ─────────────────────────────────────────────

EXCLUDE_KEYWORDS = [
    "java ", ".net developer", "python developer", "data scientist",
    "devops engineer", "ios developer", "android developer", "php developer",
    "ruby", "golang", "10 years", "10+ years", "8 years", "8+ years",
    "15 years", "20 years", "senior manager", "director", "vp of"
]

INCLUDE_KEYWORDS = [
    "angular", "typescript", "react native", "react",
    "frontend", "front-end", "front end", "ui developer",
    "javascript developer", "ionic", "expo"
]

def is_relevant(job: dict) -> bool:
    title       = job.get("title", "").lower()
    description = job.get("description", "").lower()
    combined    = title + " " + description
    has_include = any(kw in combined for kw in INCLUDE_KEYWORDS)
    has_exclude = any(kw in title for kw in EXCLUDE_KEYWORDS)
    return has_include and not has_exclude

def extract_skills(text: str) -> str:
    text_lower = text.lower()
    skill_list = [
        "Angular", "React Native", "React", "TypeScript", "JavaScript",
        "RxJS", "HTML5", "CSS3", "SCSS", "Node.js", "REST APIs",
        "Cypress", "Jest", "Selenium", "Jenkins", "CI/CD",
        "PostgreSQL", "MySQL", "MongoDB", "Ionic", "Expo",
        "Bootstrap", "Webpack", "JWT", "OAuth2", "Agile", "JIRA",
        "Git", "GitHub", "Docker", "AWS", "Azure"
    ]
    found = [s for s in skill_list if s.lower() in text_lower]
    return ", ".join(found[:10]) if found else "Frontend, TypeScript, JavaScript"

def parse_job(raw: dict) -> dict:
    company      = raw.get("company", {}).get("display_name", "Unknown Company")
    title        = raw.get("title", "")
    location     = raw.get("location", {}).get("display_name", "India")
    salary_min   = raw.get("salary_min")
    salary_max   = raw.get("salary_max")
    description  = raw.get("description", "")[:500].strip() + "..."
    redirect_url = raw.get("redirect_url", "")
    created      = raw.get("created", "")[:10] if raw.get("created") else datetime.now().strftime("%Y-%m-%d")
    job_id       = raw.get("id", "")

    # ── Resolve exact company job page URL ──────────
    print(f"   🔗 Resolving URL for: {company} – {title[:40]}...")
    exact_url = resolve_exact_url(redirect_url)

    # ── Salary ──────────────────────────────────────
    if salary_min and salary_max:
        salary = f"₹{int(salary_min):,} – ₹{int(salary_max):,}/yr"
    elif salary_min:
        salary = f"₹{int(salary_min):,}+/yr"
    else:
        salary = "Not disclosed"

    # ── Location type ────────────────────────────────
    loc_lower = location.lower()
    if "remote" in loc_lower:
        loc_type = f"{location} (Remote)"
    elif any(c in loc_lower for c in ["bangalore", "chennai", "hyderabad", "mumbai", "pune", "delhi", "gurgaon", "noida"]):
        loc_type = f"{location} (Onsite/Hybrid)"
    else:
        loc_type = location

    # ── Guess HR email ───────────────────────────────
    company_slug = company.lower().replace(" ", "").replace(",","").replace(".","")[:20]
    hr_email = f"careers@{company_slug}.com"

    return {
        "id":          job_id,
        "company":     company,
        "role":        title,
        "experience":  "0-3 Years",
        "salary":      salary,
        "location":    loc_type,
        "hr_email":    hr_email,
        "job_link":    exact_url,       # ← EXACT company URL
        "source":      "Adzuna API (Live)",
        "posted_date": created,
        "key_skills":  extract_skills(description),
        "description": description,
    }

# ─────────────────────────────────────────────
# SECTION 5: WHY I MATCH GENERATOR
# ─────────────────────────────────────────────

def generate_why_i_match(job: dict) -> str:
    text = (job.get("key_skills", "") + " " + job.get("description", "")).lower()
    highlights = []

    if any(s in text for s in ["angular", "angularjs"]):
        highlights.append("3 yrs Angular (v2–10+) across 4 enterprise projects")
    if "rxjs" in text:
        highlights.append("RxJS observables for real-time data streaming")
    if "react native" in text or "ionic" in text or "expo" in text:
        highlights.append("React Native (Expo) + Ionic mobile development")
    if "typescript" in text:
        highlights.append("TypeScript across all production projects")
    if any(s in text for s in ["cypress", "selenium", "testcafe", "playwright", "jest", "karma"]):
        highlights.append("Test automation – TestCafe/Cypress/Selenium, 30% reduction")
    if any(s in text for s in ["jenkins", "ci/cd", "github actions", "gitlab"]):
        highlights.append("Jenkins + GitLab CI/CD + GitHub Actions pipelines")
    if any(s in text for s in ["jwt", "oauth2", "auth"]):
        highlights.append("JWT + OAuth2 REST API security integration")
    if any(s in text for s in ["payment", "razorpay", "fintech"]):
        highlights.append("LLD Payment Apps cert + Razorpay knowledge")
    if any(s in text for s in ["performance", "lazy loading", "virtual scroll"]):
        highlights.append("20% performance improvement via lazy loading + OnPush")
    if any(s in text for s in ["solid", "clean code", "tdd"]):
        highlights.append("SOLID Principles & Clean Code certification (Scaler)")
    if any(s in text for s in ["real-time", "websocket", "dashboard"]):
        highlights.append("Built 6+ real-time factory dashboards with live data")
    if any(s in text for s in ["component", "library", "reusable"]):
        highlights.append("Architected 15+ reusable components, 35% faster velocity")
    if any(s in text for s in ["agile", "scrum", "sprint", "jira"]):
        highlights.append("Full Agile/Scrum – sprint planning, standups, JIRA")
    if "ai" in text:
        highlights.append("Integrated ML/AI outputs into industrial dashboards")

    if not highlights:
        highlights.append("Frontend TypeScript/Angular skills match job requirements")

    score = len(highlights)
    match_level = "🟢 HIGH MATCH" if score >= 5 else "🟡 MEDIUM MATCH" if score >= 3 else "🔴 LOW MATCH"
    return f"{match_level} | " + " | ".join(highlights[:4])

# ─────────────────────────────────────────────
# SECTION 6: DEDUPLICATION WITH HISTORY
# ─────────────────────────────────────────────

def load_seen_ids() -> set:
    if os.path.exists(HISTORY_FILE):
        with open(HISTORY_FILE, "r") as f:
            return set(json.load(f))
    return set()

def save_seen_ids(seen: set):
    with open(HISTORY_FILE, "w") as f:
        json.dump(list(seen), f)

def deduplicate(jobs: list, seen_ids: set) -> list:
    unique = []
    seen_titles = set()
    for job in jobs:
        job_id = job.get("id", "")
        key    = (job.get("company","") + job.get("role","")).lower().strip()
        if job_id not in seen_ids and key not in seen_titles:
            unique.append(job)
            seen_titles.add(key)
    return unique

# ─────────────────────────────────────────────
# SECTION 7: EXPORT CSV + EXCEL
# ─────────────────────────────────────────────

HEADERS = [
    "Company Name", "Role", "Experience", "Salary", "Location",
    "HR Email", "Job Link", "Source", "Posted Date",
    "Key Skills", "Why I Match", "Job Description"
]

def build_row(job: dict) -> dict:
    return {
        "Company Name":   job.get("company", ""),
        "Role":           job.get("role", ""),
        "Experience":     job.get("experience", "0-3 Years"),
        "Salary":         job.get("salary", "Not disclosed"),
        "Location":       job.get("location", ""),
        "HR Email":       job.get("hr_email", ""),
        "Job Link":       job.get("job_link", ""),
        "Source":         job.get("source", "Adzuna API"),
        "Posted Date":    job.get("posted_date", ""),
        "Key Skills":     job.get("key_skills", ""),
        "Why I Match":    generate_why_i_match(job),
        "Job Description": job.get("description", ""),
    }

def export_csv(rows: list, path: str):
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=HEADERS)
        writer.writeheader()
        writer.writerows(rows)
    print(f"  ✅ CSV  → {path}")

def export_xlsx(rows: list, path: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Job Listings"

    header_fill = PatternFill("solid", start_color="1F3864", end_color="1F3864")
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    thin   = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    fill_a = PatternFill("solid", start_color="EBF1F8", end_color="EBF1F8")
    fill_b = PatternFill("solid", start_color="FFFFFF", end_color="FFFFFF")

    for ci, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    col_widths = [25, 38, 12, 20, 28, 30, 50, 18, 14, 42, 50, 58]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 30

    for ri, row in enumerate(rows, 2):
        fill = fill_a if ri % 2 == 0 else fill_b
        for ci, h in enumerate(HEADERS, 1):
            cell = ws.cell(row=ri, column=ci, value=row.get(h, ""))
            cell.font = Font(name="Arial", size=9)
            cell.fill = fill
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = border
        ws.row_dimensions[ri].height = 55

    ws.freeze_panes = "A2"
    wb.save(path)
    print(f"  ✅ XLSX → {path}")

# ─────────────────────────────────────────────
# SECTION 8: MAIN
# ─────────────────────────────────────────────

def main():
    print("\n" + "="*58)
    print("  GOKUL'S DAILY JOB SEARCH — LIVE API VERSION")
    print(f"  Run Date : {datetime.now().strftime('%Y-%m-%d %H:%M')} IST")
    print("="*58)

    if not APP_ID or not APP_KEY:
        print("\n❌ ERROR: ADZUNA_APP_ID or ADZUNA_APP_KEY not set!")
        print("   Add them as GitHub Secrets and reference in yml env: block.")
        return

    seen_ids = load_seen_ids()
    print(f"\n📂 Previously seen jobs: {len(seen_ids)}")

    print("\n🔍 Fetching live jobs from Adzuna API...")
    all_raw = []
    for query in SEARCH_QUERIES:
        print(f"   Searching: '{query}'...")
        results = fetch_jobs(query, pages=2)
        all_raw.extend(results)
        print(f"   → {len(results)} results")

    print(f"\n📋 Total raw results : {len(all_raw)}")

    relevant = [r for r in all_raw if is_relevant(r)]
    print(f"✅ After filter      : {len(relevant)}")

    print("\n🔗 Resolving exact job URLs...")
    parsed   = [parse_job(r) for r in relevant]

    new_jobs = deduplicate(parsed, seen_ids)
    print(f"\n🆕 New jobs today    : {len(new_jobs)}")

    if not new_jobs:
        print("\n⚠️  No new jobs found today. Try again tomorrow!")
        return

    rows   = [build_row(j) for j in new_jobs]
    high   = sum(1 for r in rows if "HIGH"   in r["Why I Match"])
    medium = sum(1 for r in rows if "MEDIUM" in r["Why I Match"])
    low    = sum(1 for r in rows if "LOW"    in r["Why I Match"])

    print(f"\n🎯 Match Summary:")
    print(f"   🟢 High Match   : {high}")
    print(f"   🟡 Medium Match : {medium}")
    print(f"   🔴 Low Match    : {low}")

    folder = "job_listing"
  if os.path.exists(folder):
    today = datetime.now().strftime("%Y-%m-%d")

    csv_path = os.path.join(folder, f"job_listing_{today}.csv")
    xlsx_path = os.path.join(folder, f"job_listing_{today}.xlsx")

    print("\n💾 Saving files...")
    export_csv(rows, csv_path)
    export_xlsx(rows, xlsx_path)

    new_ids = {j.get("id") for j in new_jobs if j.get("id")}
    save_seen_ids(seen_ids | new_ids)
    print(f"📝 History updated : {len(new_ids)} new IDs saved")

    try:
        from google.colab import files
        print("\n📥 Downloading files...")
        files.download(csv_path)
        files.download(xlsx_path)
    except ImportError:
        print(f"\n📂 Files saved in: {os.path.abspath('.')}")

    print("\n" + "="*58)
    print(f"  ✅ DONE! {len(new_jobs)} fresh jobs with exact URLs exported.")
    print("="*58 + "\n")


if __name__ == "__main__":
    main()
